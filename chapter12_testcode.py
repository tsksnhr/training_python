# import module
import os
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# read workbook object
test = os.getcwd()
print(test)

path = "example.xlsx"
if (os.path.exists(path)):
    wb = openpyxl.load_workbook("example.xlsx")
    print(type(wb))
else:
    print("No path exists.")

# get sheet-name-list
sheet_list = wb.sheetnames
print(sheet_list)
print()

# get sheet-name
sheet = wb['Sheet3']
print(sheet)
print(type(sheet))
print(sheet.title)
print()

# get active-sheet
active_sheet = wb.active
print(active_sheet)
print()

# get cell
cell1 = active_sheet["A1"]
print(cell1.value)
print(cell1.row)
print(cell1.column)
print(cell1.coordinate)
print()

# get cell
cell2 = active_sheet["B1"]
print(cell2.value)
print(cell2.row)
print(cell2.column)
print(cell2.coordinate)
print()

# get cell by cell-method
c = active_sheet.cell(row = 1, column = 2)
print(c.coordinate)
print(c.value)

for i in range(1, 8, 2):
    print(i, active_sheet.cell(row = i, column = 2).value)
print()

# get cheet-zize
max_r = active_sheet.max_row
max_c = active_sheet.max_column
print(max_r, max_c)
print()

# exchange string and figure
print(get_column_letter(5))
print(get_column_letter(38))
print(column_index_from_string("C"))
print(column_index_from_string("ABC"))

print(get_column_letter(active_sheet.max_column))
print()

# get rows and columns
t = tuple(active_sheet["A1":"C3"])
print(t)
print()

for row_of_cell_obj in active_sheet["A1":"C3"]:
    for cell_obj in row_of_cell_obj:
        print(cell_obj.coordinate, cell_obj.value)
    print("--- END OF ROW ---")
